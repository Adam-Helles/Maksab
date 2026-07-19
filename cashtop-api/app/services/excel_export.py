"""
خدمة تصدير Excel — تولّد workbook متعدد الأوراق

⚠️ إصلاح أمني جوهري: كل الأوراق (مبيعات، مخزون، ديون عملاء، ديون
موردين) كانت بتصدّر بيانات كل التجار المسجلين بالنظام مع بعض بدون أي
فلترة. هاد يعني أي تاجر يصدّر تقرير Excel كان ممكن (حسب الصلاحيات)
يشوف أرقام مبيعات ومخزون وديون تجار تانيين. صار store_id إجباري
بكل دالة.
"""
import io
from datetime import date, datetime
from typing import Optional
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")
TOTAL_FILL   = PatternFill("solid", fgColor="E8F5E9")
TOTAL_FONT   = Font(bold=True, color="1B5E20")
CENTER       = Alignment(horizontal="center", vertical="center")
RIGHT        = Alignment(horizontal="right")
THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, cols: int, alt: bool = False):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        if alt:
            cell.fill = ALT_FILL
        cell.border = THIN_BORDER
        cell.alignment = RIGHT


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(max_len + 4, max_w))


# ══════════════════════════════════════════════════════════
#  ورقة المبيعات
# ══════════════════════════════════════════════════════════

def _sales_sheet(ws, db: Session, store_id: int, date_from=None, date_to=None):
    from app.models.invoice import Invoice, InvoiceType, InvoiceStatus
    from sqlalchemy import func

    ws.title = "المبيعات"
    ws.sheet_view.rightToLeft = True

    headers = ["رقم الفاتورة", "التاريخ", "العميل", "الأصناف",
            "المجموع", "الخصم", "الضريبة", "الإجمالي", "المدفوع", "المتبقي", "الحالة"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    q = db.query(Invoice).filter(
        Invoice.store_id == store_id,
        Invoice.invoice_type == InvoiceType.SALE,
        Invoice.status == InvoiceStatus.COMPLETED,
    )
    if date_from:
        q = q.filter(func.date(Invoice.created_at) >= date_from)
    if date_to:
        q = q.filter(func.date(Invoice.created_at) <= date_to)
    invoices = q.order_by(Invoice.created_at.desc()).all()

    total_sales = total_profit = 0.0
    for row_i, inv in enumerate(invoices, 2):
        customer_name = inv.customer.name if inv.customer else "نقدي"
        status_ar = {"paid": "مدفوعة", "partial": "جزئي", "unpaid": "آجل"}.get(inv.payment_status.value, inv.payment_status.value)

        ws.cell(row=row_i, column=1,  value=inv.invoice_number)
        ws.cell(row=row_i, column=2,  value=inv.created_at.strftime("%Y-%m-%d"))
        ws.cell(row=row_i, column=3,  value=customer_name)
        ws.cell(row=row_i, column=4,  value=len(inv.items))
        ws.cell(row=row_i, column=5,  value=inv.subtotal)
        ws.cell(row=row_i, column=6,  value=inv.discount_amount + (inv.subtotal * inv.discount_percent / 100))
        ws.cell(row=row_i, column=7,  value=inv.tax_amount)
        ws.cell(row=row_i, column=8,  value=inv.total)
        ws.cell(row=row_i, column=9,  value=inv.paid_amount)
        ws.cell(row=row_i, column=10, value=inv.remaining_amount)
        ws.cell(row=row_i, column=11, value=status_ar)
        _style_data_row(ws, row_i, len(headers), alt=(row_i % 2 == 0))
        total_sales += inv.total

    total_row = len(invoices) + 2
    ws.cell(row=total_row, column=7, value="الإجمالي")
    ws.cell(row=total_row, column=8, value=round(total_sales, 2))
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT

    _auto_width(ws)


# ══════════════════════════════════════════════════════════
#  ورقة المخزون
# ══════════════════════════════════════════════════════════

def _inventory_sheet(ws, db: Session, store_id: int):
    from app.models.product import Product

    ws.title = "المخزون"
    ws.sheet_view.rightToLeft = True

    headers = ["الاسم", "الباركود", "الصنف", "المخزون (قطعة)",
            "المخزون (كرتونة)", "سعر التكلفة", "سعر التجزئة",
            "قيمة المخزون", "هامش الربح %", "الحد الأدنى", "الحالة"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    products = db.query(Product).filter(
        Product.store_id == store_id,
        Product.is_active == True, Product.is_deleted == False,
    ).order_by(Product.name).all()

    total_value = 0.0
    for row_i, p in enumerate(products, 2):
        stock_val = p.stock_quantity * p.cost_price
        status    = "⚠️ منخفض" if p.is_low_stock else "✅ طبيعي"
        cat_name  = p.category.name if p.category else "—"

        ws.cell(row=row_i, column=1,  value=p.name)
        ws.cell(row=row_i, column=2,  value=p.barcode_piece or "—")
        ws.cell(row=row_i, column=3,  value=cat_name)
        ws.cell(row=row_i, column=4,  value=p.stock_quantity)
        ws.cell(row=row_i, column=5,  value=round(p.stock_in_cartons, 2))
        ws.cell(row=row_i, column=6,  value=p.cost_price)
        ws.cell(row=row_i, column=7,  value=p.retail_price)
        ws.cell(row=row_i, column=8,  value=round(stock_val, 2))
        ws.cell(row=row_i, column=9,  value=p.profit_margin)
        ws.cell(row=row_i, column=10, value=p.min_stock_alert)
        ws.cell(row=row_i, column=11, value=status)
        _style_data_row(ws, row_i, len(headers), alt=(row_i % 2 == 0))
        total_value += stock_val

    total_row = len(products) + 2
    ws.cell(row=total_row, column=7,  value="قيمة المخزون الكلية")
    ws.cell(row=total_row, column=8,  value=round(total_value, 2))
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
    _auto_width(ws)


# ══════════════════════════════════════════════════════════
#  ورقة ديون العملاء
# ══════════════════════════════════════════════════════════

def _customer_debts_sheet(ws, db: Session, store_id: int):
    from app.models.customer import Customer

    ws.title = "ديون العملاء"
    ws.sheet_view.rightToLeft = True

    headers = ["اسم العميل", "الجوال", "الدين الحالي", "حد الائتمان", "الرصيد المتاح", "الحالة"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    customers = db.query(Customer).filter(
        Customer.store_id == store_id,
        Customer.current_debt > 0,
        Customer.is_deleted == False,
    ).order_by(Customer.current_debt.desc()).all()

    total_debt = 0.0
    for row_i, c in enumerate(customers, 2):
        status = "⛔ تجاوز الحد" if not c.can_buy_on_credit else "✅ ضمن الحد"
        ws.cell(row=row_i, column=1, value=c.name)
        ws.cell(row=row_i, column=2, value=c.phone or "—")
        ws.cell(row=row_i, column=3, value=c.current_debt)
        ws.cell(row=row_i, column=4, value=c.credit_limit if c.credit_limit > 0 else "بلا حد")
        ws.cell(row=row_i, column=5, value=round(c.available_credit, 2) if c.credit_limit > 0 else "∞")
        ws.cell(row=row_i, column=6, value=status)
        _style_data_row(ws, row_i, len(headers), alt=(row_i % 2 == 0))
        total_debt += c.current_debt

    total_row = len(customers) + 2
    ws.cell(row=total_row, column=2, value="إجمالي الديون")
    ws.cell(row=total_row, column=3, value=round(total_debt, 2))
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
    _auto_width(ws)


# ══════════════════════════════════════════════════════════
#  ورقة ديون الموردين
# ══════════════════════════════════════════════════════════

def _supplier_debts_sheet(ws, db: Session, store_id: int):
    from app.models.supplier import Supplier

    ws.title = "ديون الموردين"
    ws.sheet_view.rightToLeft = True

    headers = ["اسم المورد", "الشركة", "الجوال", "المبلغ المستحق"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    suppliers = db.query(Supplier).filter(
        Supplier.store_id == store_id,
        Supplier.balance > 0,
        Supplier.is_deleted == False,
    ).order_by(Supplier.balance.desc()).all()

    total = 0.0
    for row_i, s in enumerate(suppliers, 2):
        ws.cell(row=row_i, column=1, value=s.name)
        ws.cell(row=row_i, column=2, value=s.company or "—")
        ws.cell(row=row_i, column=3, value=s.phone or "—")
        ws.cell(row=row_i, column=4, value=s.balance)
        _style_data_row(ws, row_i, len(headers), alt=(row_i % 2 == 0))
        total += s.balance

    total_row = len(suppliers) + 2
    ws.cell(row=total_row, column=3, value="الإجمالي")
    ws.cell(row=total_row, column=4, value=round(total, 2))
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
    _auto_width(ws)


# ══════════════════════════════════════════════════════════
#  الدالة الرئيسية — Full Workbook
# ══════════════════════════════════════════════════════════

def generate_excel_report(
    db: Session,
    store_id: int,               # ⚠️ جديد — إجباري، أول باراميتر بعد db
    report_type: str = "full",
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if report_type in ("full", "sales"):
        _sales_sheet(wb.create_sheet(), db, store_id, date_from, date_to)

    if report_type in ("full", "inventory"):
        _inventory_sheet(wb.create_sheet(), db, store_id)

    if report_type in ("full", "debts"):
        _customer_debts_sheet(wb.create_sheet(), db, store_id)
        _supplier_debts_sheet(wb.create_sheet(), db, store_id)

    # حماية: openpyxl ترفض حفظ workbook بدون أوراق
    if not wb.sheetnames:
        ws = wb.create_sheet("لا توجد بيانات")
        ws["A1"] = "لا توجد بيانات لعرضها في هذا التقرير"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()